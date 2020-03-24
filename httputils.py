# -*- coding: <utf-8> -*-
import urllib.request
from bs4 import BeautifulSoup
import openpyxl
# params=urllib.parse.urlencode({'pageNo': 4})
# req=urllib.request.Request("http://wsjkw.cq.gov.cn/yqxxyqtb/index.jhtml",data=params.encode("utf-8") ,method="POST")
# response=urllib.request.urlopen(req)
# if(response.status==200):  
#   data=response.read()
#   soup = BeautifulSoup(data.decode(encoding="utf-8"), 'html.parser')
#   for el in soup.select("body .mdiv .newslist ul li"):
#     req1=urllib.request.Request("http://wsjkw.cq.gov.cn"+el.a.attrs['href'],data=None,method="GET")
#     response1=urllib.request.urlopen(req1)
#     data1=response1.read().decode(encoding="utf-8")
#     soup1 = BeautifulSoup(data1, 'html.parser')
#     title=soup1.select("body .combox .lfbox .news_top .news_tit")[0].string
#     if(title.find("重庆市新型冠状病毒感染的肺炎疫情情况")>0):
#         print("---------{}----------".format(title))
#         for p in soup1.select("body .combox .lfbox .uedit p"): 
#          if(p.string and p.string.startswith("确诊病例中")):
#             print(p.string) 

#获取重庆卫键委数据
req=urllib.request.Request("http://wsjkw.cq.gov.cn/yqxxtzgg/",method="GET")
response=urllib.request.urlopen(req)
if(response.status==200):  
  data=response.read()
  soup = BeautifulSoup(data.decode(encoding="utf-8"), 'html.parser')
  resultData=[]
  for el in soup.select("body .mdiv .newslist ul li"):
    resultData.append((el.a.select("span")[0].string,str(el.a.select("span")[1].string).strip(),"http://wsjkw.cq.gov.cn"+el.a.attrs["href"]))
    #导出到excel
    workbook=openpyxl.workbook.Workbook()
    worksheet=workbook.worksheets[0]
    worksheet.cell(1,1,value="title")
    worksheet.cell(1,2,value="date")
    worksheet.cell(1,3,value="url")
    rowIndex=1
    for (d,t,u) in resultData:
        rowIndex=rowIndex+1
        worksheet.cell(rowIndex,1,value=t)
        worksheet.cell(rowIndex,2,value=d)
        worksheet.cell(rowIndex,3,value=u)
    workbook.save("E:\\project\\yiqing\\data\\2020.xlsx")
    