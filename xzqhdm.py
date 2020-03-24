# -*- coding: <utf-8> -*-
import urllib.request
from bs4 import BeautifulSoup
import logging
import openpyxl
import time
class Xzqdm:
  __baseUrl=""
  def __init__(self,baseUrl):
     self.__baseUrl=baseUrl
  def getXzqdm(self,xzqdmData,url):
        time.sleep(1)
        if(not url):
          return xzqdmData
        print(url)
        count=0
        isSuccess=False
        while(not isSuccess and count<5):
          count=count+1
          try:
            req=urllib.request.Request(url,method="GET")      
            response=urllib.request.urlopen(req)
            isSuccess=True
          except urllib.error.HTTPError:
            print()
        if(not isSuccess):
          logging.error("{},执行错误".format(url))
        if(response.status==200):  
          data=response.read()
          soup = BeautifulSoup(data.decode(encoding="gbk"), 'html.parser')
          index =0
          for tr in soup.select("body table")[1].select("table table table table tr"):         
            if(index!=0):
              if(tr.select("td")[0].a):
                xzqdmData[tr.select("td")[1].a.string]=tr.select("td")[0].a.string          
                self.getXzqdm(xzqdmData,url[0:url.rfind("/")+1]+tr.select("td")[0].a.attrs['href'])
              else:
                l=len(tr.select("td"))
                xzqdmData[tr.select("td")[l-1].string]=tr.select("td")[0].string          
                self.getXzqdm(xzqdmData,None)
            index =index+1
  def writeExcel(self,name,xzqdmData):
      #导出到excel
      workbook=openpyxl.workbook.Workbook()
      worksheet=workbook.create_sheet(name,0)
      rowIndex=0
      for k,v in xzqdmData.items():
        rowIndex=rowIndex+1
        worksheet.cell(rowIndex,1,value=k)
        worksheet.cell(rowIndex,2,value=v)
      workbook.save("D:\\行政区代码\\"+name+".xlsx")

if __name__ == "__main__":
  xz=["北京市","天津市","河北省"]
  logging.basicConfig(level=logging.DEBUG,filename='new.log', filemode='w')
  baseUrl="http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2018/"
  xzqdm= Xzqdm(baseUrl)
  xzqdmData={}
  # xzqdm.getXzqdm(xzqdmData,'http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2018/11/01/110101.html')
  req=urllib.request.Request(baseUrl+"index.html",method="GET")
  response=urllib.request.urlopen(req)
  if(response.status==200):  
    data=response.read()
    soup = BeautifulSoup(data.decode(encoding="gbk"), 'html.parser')
    for el in soup.select(".provincetr"):
      for td in el.children:
        if td.a.contents[0] not in xz:
          xzqdmData={}
          xzqdmData[td.a.contents[0]]="" 
          xzqdm.getXzqdm(xzqdmData,baseUrl+td.a.attrs['href'])
          xzqdm.writeExcel(td.a.contents[0],xzqdmData)
  print("完成")