# -*- coding: <utf-8> -*-
import urllib.request
from bs4 import BeautifulSoup
import openpyxl
import datetime

#获取重庆卫键委数据
# req=urllib.request.Request("http://wsjkw.cq.gov.cn/yqxxtzgg/",method="GET")
# response=urllib.request.urlopen(req)
# if(response.status==200):  
#   data=response.read()
#   soup = BeautifulSoup(data.decode(encoding="utf-8"), 'html.parser')
#   resultData=[]
#   for el in soup.select("body .mdiv .newslist ul li"):
#     resultData.append((el.a.select("span")[0].string,str(el.a.select("span")[1].string).strip(),"http://wsjkw.cq.gov.cn"+el.a.attrs["href"]))
#     #导出到excel
#     workbook=openpyxl.workbook.Workbook()
#     worksheet=workbook.worksheets[0]
#     worksheet.cell(1,1,value="title")
#     worksheet.cell(1,2,value="date")
#     worksheet.cell(1,3,value="url")
#     rowIndex=1
#     for (d,t,u) in resultData:
#         rowIndex=rowIndex+1
#         worksheet.cell(rowIndex,1,value=t)
#         worksheet.cell(rowIndex,2,value=d)
#         worksheet.cell(rowIndex,3,value=u)
#     workbook.save("E:\\project\\yiqing\\data\\2020.xlsx")


#更新各区县，每日数据
dd=datetime.date(2020,2,20)
workbook=openpyxl.open("E:\\project\\yiqing\\data\\"+dd.strftime("%Y%m%d")+".xlsx")
sheet=workbook.worksheets[0]
data={}
for i in range(4,sheet.max_row):
    key=sheet.cell(i,2).value
    if(len(key)==3):
      key=key[0:2]
    if(sheet.cell(i,4).value):
      data[key]=int(str(sheet.cell(i,4).value))
    else:
      data[key]=0
workbook1=openpyxl.open("E:\\project\\yiqing\\data\\重庆各区县确诊历史数据.xlsx")
sheet1=workbook1.worksheets[0]
col=sheet1.max_column+1
sheet1.cell(1,col,value=dd.strftime(r"%#m月%#d日"))
for i in range(2,sheet1.max_row):
    if(sheet1.cell(i,1).value in data):
      value=data[sheet1.cell(i,1).value]
      sheet1.cell(i,col,value=value)
workbook1.save("E:\\project\\yiqing\\data\\重庆各区县确诊历史数据1.xlsx")
    