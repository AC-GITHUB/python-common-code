# -*- coding: <utf-8> -*-
import openpyxl
import json
workbook=openpyxl.open("E:\\project\\yiqing\\data\\重庆各区县确诊历史数据.xlsx")
sheet=workbook["worksheet"]
key=[]
result=[]
for i in range(1,sheet.max_column):   
    if(sheet.cell(1,i).is_date):
      key.append(sheet.cell(1,i).value.strftime("%m月%d日"))
    else:
      key.append(sheet.cell(1,i).value)
for i in range(2,sheet.max_row):
    obj={}
    for j in range(1,sheet.max_column):
        if(str(sheet.cell(i,j).value).isdigit()):
          obj[key[j-1]]=int(sheet.cell(i,j).value)
        else:
           obj[key[j-1]]=sheet.cell(i,j).value
    result.append(obj)
print(json.dump(result,open(r"E:\test.json",mode="w"),ensure_ascii=False))