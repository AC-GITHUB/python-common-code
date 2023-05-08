# -*- coding: <utf-8> -*-
import urllib.request
from bs4 import BeautifulSoup
import logging
import openpyxl
import time


class Xzqdm:
    __base_url = ""
    __count = 0

    def __init__(self, base_url):
        self.__base_url = base_url

    def get_xzqdm(self, xzqdm_data, url):
        if (not url):
            return xzqdm_data
        print(url)
        if (self.__count % 1000 == 0):
            time.sleep(3)
        self.__count = self.__count + 1
        count = 0
        is_success = False
        while (not is_success and count < 5):
            time.sleep(0.01)
            count = count + 1
            try:
                req = urllib.request.Request(url, method="GET")
                response = urllib.request.urlopen(req, timeout=10)
                is_success = True
            except urllib.error.HTTPError as e:
                print(e)
            except Exception as e:
                print(e)
        if (not is_success):
            logging.error("{},执行错误".format(url))
        if (response.status == 200):
            data = response.read()
            soup = BeautifulSoup(data.decode(encoding="utf-8"), 'html.parser')
            index = 0
            for tr in soup.select("body table")[1].select(
                    "table table table table tr"):
                if (index != 0):
                    if (tr.select("td")[0].a):
                        xzqdm_data.append({
                            'name': tr.select("td")[1].a.string,
                            'code': tr.select("td")[0].a.string,
                            'text': str(tr)
                        })
                        self.get_xzqdm(
                            xzqdm_data, url[0:url.rfind("/") + 1] +
                            tr.select("td")[0].a.attrs['href'])
                    else:
                        le = len(tr.select("td"))
                        xzqdm_data.append({
                            'name': tr.select("td")[le - 1].string,
                            'code': tr.select("td")[0].string,
                            'text': str(tr)
                        })
                        self.get_xzqdm(xzqdm_data, None)
                index = index + 1

    def write_excel(self, name, xzqdm_data):
        # 导出到excel
        workbook = openpyxl.workbook.Workbook()
        worksheet = workbook.create_sheet(name, 0)
        row_index = 0
        for item in xzqdm_data:
            row_index = row_index + 1
            worksheet.cell(row_index, 1, value=item['name'])
            worksheet.cell(row_index, 2, value=item['code'])
            worksheet.cell(row_index, 3, value=item['text'])
        workbook.save("/Users/acc/gitworkspace/xzqdm/" + name + ".xlsx")


if __name__ == "__main__":
    xz = [
        "北京市", "天津市", "河北省", "山西省", "内蒙古自治区", "辽宁省", "吉林省", "黑龙江省", "上海市",
        "江苏省", "浙江省", "安徽省", "福建省", "江西省", "山东省", "河南省", "湖北省", "湖南省", "广东省",
        "广西壮族自治区", "海南省", "重庆市"
    ]
    logging.basicConfig(level=logging.DEBUG, filename='new.log', filemode='w')
    base_url = "http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2021/"
    xzqdm = Xzqdm(base_url)
    xzqdm_data = []
    # xzqdm.getXzqdm(xzqdmData,'http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2018/11/01/110101.html')
    req = urllib.request.Request(base_url + "index.html", method="GET")
    response = urllib.request.urlopen(req)
    if (response.status == 200):
        data = response.read()
        soup = BeautifulSoup(data.decode(encoding="utf-8"), 'html.parser')
        for el in soup.select(".provincetr"):
            for td in el.children:
                if td.a.contents[0] not in xz:
                    xzqdm_data = []
                    #   xzqdmData[td.a.contents[0]]=""
                    xzqdm.get_xzqdm(xzqdm_data, base_url + td.a.attrs['href'])
                    xzqdm.write_excel(td.a.contents[0], xzqdm_data)
    print("完成")
