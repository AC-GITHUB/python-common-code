# -*- coding: <utf-8> -*-
import openpyxl
import json


class ExcelUtils(object):

    def __init__(self):
        super()

    def excel_to_json(self, excel_file, sheet_name):
        workbook = openpyxl.open(excel_file)
        sheet = workbook[sheet_name]
        key = []
        result = []
        for i in range(1, sheet.max_column):
            if (sheet.cell(1, i).is_date):
                key.append(sheet.cell(1, i).value.strftime("%m月%d日"))
            else:
                key.append(sheet.cell(1, i).value)
        for i in range(2, sheet.max_row):
            obj = {}
            for j in range(1, sheet.max_column):
                if (str(sheet.cell(i, j).value).isdigit()):
                    obj[key[j - 1]] = int(sheet.cell(i, j).value)
                else:
                    obj[key[j - 1]] = sheet.cell(i, j).value
            result.append(obj)
            return json.dump(result, ensure_ascii=False)
