# -*- coding: <utf-8> -*-
import openpyxl
import mysql.connector


class MysqlUtils(object):
    cnx = None

    def __init__(self):
        super()

    def init_connection(self, user, password, host, port, db):
        self.cnx = mysql.connector.connection.MySQLConnection(user=user,
                                                              password=password,
                                                              host=host,
                                                              port=port,
                                                              database=db)

    def export_to_excel(self, sql, excel_filename):
        cursor = self.cnx.cursor()
        workbook = openpyxl.workbook.Workbook()
        worksheet = workbook.create_sheet("sheet1", 0)
        cursor.execute(sql)
        col_index = 0
        for row in cursor:
            col_index = col_index + 1
            worksheet.cell(1, col_index, value=row[1])
        workbook.save(excel_filename)
