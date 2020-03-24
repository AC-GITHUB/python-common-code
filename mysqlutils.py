# -*- coding: <utf-8> -*-
import openpyxl
import mysql.connector
#创建连接
# cnx = mysql.connector.connection.MySQLConnection(user='root', password='eversec123098',host='192.168.205.232',database='information_schema',port=13306)
# cursor=cnx.cursor()
#读取数据
# cursor.execute("SELECT * FROM `COLUMNS` WHERE TABLE_SCHEMA='asset' and TABLE_NAME='asset_attribute'")
# data=None 
# for row in cursor:
#   print(row)
#   data=row
# 插入数据
# insertSql = (
#   "INSERT INTO children_onlinerate_daily (date, parent_sum_count, children_sum_count, day_nine_am_online_count,day_nine_am_online_rate,day_twelve_am_online_count,day_twelve_am_online_rate,day_nine_pm_online_count,day_nine_pm_online_rate) "
#   "VALUES (%s, %s, %s, %s,%s,%s,%s,%s,%s)")
# print(data)
# print(cursor.execute(insertSql, data))
# cnx.commit()

#导出到excel
# workbook=openpyxl.workbook.Workbook()
# worksheet=workbook.create_sheet("mysql",0)
# cursor.execute("select * from children_onlinerate_daily")
# rowIndex=1
# colIndex=0
# for name in cursor.column_names:
#   colIndex=colIndex+1
#   worksheet.cell(1,colIndex,value=name)
# for row in cursor:
#   rowIndex=rowIndex+1
#   colIndex=0
#   for val in row:
#     colIndex=colIndex+1
#     worksheet.cell(rowIndex,colIndex,value=val)
# workbook.save(r"E:\mytest.xlsx")
# workbook=openpyxl.load_workbook(r"E:\mytest.xlsx")
# worksheet=workbook.worksheets[0]
# print(workbook.sheetnames)
# print(worksheet.title)

cnx = mysql.connector.connection.MySQLConnection(user='root', password='eversec123098',host='192.168.205.232',database='information_schema',port=13306)
cursor=cnx.cursor()
#读取数据
# cursor.execute("SELECT COLUMN_NAME,COLUMN_COMMENT FROM `COLUMNS` WHERE TABLE_SCHEMA='asset' and TABLE_NAME='asset_task'")
# for row in cursor:
#   name=row[0]
#   l=len(row[0])
#   v=name[0]
#   for i in range(1,l,1):
#     if(name[i]=="_"):
#       v=v
#     elif(name[i-1]=="_"):
#       v+=name[i].upper()
#     else:
#       v+=name[i]
#   print("@ApiImplicitParam(name = \"{}\", value = \"{}\", paramType = ParamType.QUERY),".format(v,row[1]))
workbook=openpyxl.workbook.Workbook()
worksheet=workbook.create_sheet("资产",0)
cursor.execute("SELECT COLUMN_NAME,COLUMN_COMMENT FROM `COLUMNS` WHERE TABLE_SCHEMA='asset' and TABLE_NAME='asset_attribute'")
colIndex=0
for row in cursor:
  colIndex=colIndex+1
  worksheet.cell(1,colIndex,value=row[1])
workbook.save(r"E:\mytest.xlsx")